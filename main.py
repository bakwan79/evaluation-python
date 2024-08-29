import tkinter as tk
from tkinter import messagebox, ttk
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter  # Importation pour faciliter la gestion des colonnes Excel
from datetime import datetime  # (changement) Importation de datetime
import string  # Importer string pour gérer la ponctuation

# Chemins par défaut vers les répertoires des évaluations et des résultats
evaluations_dir = '/home/brice/evals/evaluations/'
results_dir = '/home/brice/evals/results/'
completed_evaluations_file = '/home/brice/evals/completed_evaluations.json'

class QuizApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QCM")
        self.evaluations = []
        self.student_name = ""
        self.current_question_index = 0
        self.score = 0
        self.completed_evaluations = set()
        self.radio_buttons_frame = None
        self.open_answer_frame = None
        self.selected_option = tk.StringVar() 
        self.create_widgets()
          
    def create_widgets(self):
        """Crée l'interface graphique."""
        self.name_label = tk.Label(self.root, text="Veuillez entrer votre nom:")
        self.name_label.pack(pady=10)

        self.name_entry = tk.Entry(self.root)
        self.name_entry.pack(pady=5)

        self.validate_name_button = tk.Button(self.root, text="Valider le nom", command=self.validate_name)
        self.validate_name_button.pack(pady=10)

        self.evaluation_label = tk.Label(self.root, text="Sélectionnez une évaluation:")
        self.evaluation_label.pack(pady=10)

        self.evaluation_var = tk.StringVar()
        self.evaluation_menu = ttk.Combobox(self.root, textvariable=self.evaluation_var, state="readonly")
        self.evaluation_menu['values'] = []  # Initialement vide
        self.evaluation_menu.pack(pady=5)
        
        # Sélectionner automatiquement la première évaluation
        self.evaluation_menu.bind('<<ComboboxSelected>>', self.on_evaluation_selected)

        self.start_button = tk.Button(self.root, text="Démarrer l'évaluation", command=self.start_quiz)
        self.start_button.pack(pady=10)
        
        # Label pour afficher la question
        self.question_label = tk.Label(self.root, text="")
        self.question_label.pack(pady=10)

        # Cadre unique pour les options de réponse (radio buttons ou open answer)
        self.answer_frame = tk.Frame(self.root)
        self.answer_frame.pack(pady=10)

        # Variable pour suivre la sélection actuelle (radio buttons ou texte libre)
        self.selected_option = tk.StringVar()  # Pour les boutons radio
        self.open_answer = tk.StringVar()  # Pour les questions ouvertes

        # Liste pour stocker les boutons radio
        self.option_buttons = []

        for option in "ABCD":
            button = tk.Radiobutton(self.answer_frame, text=option, variable=self.selected_option, value=option, state=tk.DISABLED)
            button.pack(anchor='w')
            self.option_buttons.append(button)

        # Champ de réponse pour les questions ouvertes
        self.open_answer_label = tk.Label(self.answer_frame, text="Votre réponse:")
        self.open_answer_label.pack(side=tk.LEFT, padx=5)
        self.open_answer_label.pack_forget()  # Caché par défaut

        self.open_answer_entry = tk.Entry(self.answer_frame, textvariable=self.open_answer, state=tk.DISABLED)
        self.open_answer_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.open_answer_entry.pack_forget()  # Caché par défaut

        self.submit_button = tk.Button(self.root, text="Soumettre", command=self.submit_answer, state=tk.DISABLED)
        self.submit_button.pack(pady=10)


    def on_evaluation_selected(self, event):
        """Met à jour l'évaluation sélectionnée lorsque l'utilisateur change la sélection dans la combobox."""
        selected_title = self.evaluation_var.get()
        if selected_title:
            selected_eval = self.evaluation_mapping.get(selected_title)
            if selected_eval:
                self.load_evaluation(selected_eval['file'], selected_eval['id'])

    
    def start_quiz(self):
        """Démarre l'évaluation sélectionnée."""
        self.student_name = self.name_entry.get()
        
        if not self.student_name:
            messagebox.showwarning("Attention", "Veuillez entrer votre nom.")
            return
        
        self.completed_evaluations = self.load_completed_evaluations()
        selected_title = self.evaluation_var.get()

        if not selected_title:
            messagebox.showwarning("Attention", "Veuillez sélectionner une évaluation.")
            return
        
        # Récupérer le fichier source et l'ID de l'évaluation sélectionnée
        selected_eval = self.evaluation_mapping.get(selected_title)
        
        if not selected_eval:
            messagebox.showerror("Erreur", "Évaluation sélectionnée non valide.")
            return

        # Charger l'évaluation spécifique
        self.load_evaluation(selected_eval['file'], selected_eval['id'])
        self.current_question_index = 0
        self.score = 0
        
        for button in self.option_buttons:
            button.config(state=tk.NORMAL)
        self.submit_button.config(state=tk.NORMAL)

        self.show_question()

    def validate_name(self):
        """Valide le nom de l'élève et charge les évaluations disponibles."""
        self.student_name = self.name_entry.get().strip().lower()  # Convertir en minuscules

        if not self.student_name:
            messagebox.showwarning("Attention", "Veuillez entrer votre nom.")
            return

        self.completed_evaluations = self.load_completed_evaluations()
        self.load_available_evaluations()
        self.start_button.config(state=tk.NORMAL)


    def load_completed_evaluations(self):
        """Charge les évaluations déjà passées depuis le fichier JSON."""
        if os.path.exists(completed_evaluations_file):
            with open(completed_evaluations_file, 'r') as f:
                all_evaluations = json.load(f)
            # Convertir tous les noms en minuscules
            return set(all_evaluations.get(self.student_name, []))
        return set()


    def load_available_evaluations(self):
        """Charge les évaluations non complétées dans la liste déroulante."""
        all_evaluations = []

        # Dictionnaire pour mapper les titres d'évaluations à leur fichier source et ID
        self.evaluation_mapping = {}

        for filename in os.listdir(evaluations_dir):
            if filename.endswith('.json'):
                with open(os.path.join(evaluations_dir, filename), 'r') as f:
                    evaluations = json.load(f)

                    # Si le fichier contient une seule évaluation
                    if isinstance(evaluations, dict) and 'id' in evaluations:
                        if evaluations['id'] not in self.completed_evaluations:
                            title = evaluations['title']
                            all_evaluations.append(title)
                            self.evaluation_mapping[title] = {'file': filename, 'id': evaluations['id']}

                    # Si le fichier contient plusieurs évaluations
                    elif isinstance(evaluations, list):
                        for eval_data in evaluations:
                            if eval_data['id'] not in self.completed_evaluations:
                                title = eval_data['title']
                                all_evaluations.append(title)
                                self.evaluation_mapping[title] = {'file': filename, 'id': eval_data['id']}
        
        self.evaluations = all_evaluations

        if not self.evaluations:
            messagebox.showinfo("Info", "Aucune évaluation disponible à passer.")
            self.root.quit()
        else:
            self.evaluation_menu['values'] = self.evaluations
            self.evaluation_menu.current(0)
            self.evaluation_var.set(self.evaluations[0])

    def load_evaluation(self, evaluation_file, evaluation_id):
        """Charge l'évaluation sélectionnée."""
        with open(os.path.join(evaluations_dir, evaluation_file), 'r') as ef:
            evaluations = json.load(ef)
            
            # Si le fichier contient une seule évaluation
            if isinstance(evaluations, dict) and evaluations['id'] == evaluation_id:
                self.evaluation = evaluations
            # Si le fichier contient plusieurs évaluations
            elif isinstance(evaluations, list):
                for eval_data in evaluations:
                    if eval_data['id'] == evaluation_id:
                        self.evaluation = eval_data
                        break

    
    def show_question(self):
        """Affiche la question actuelle."""
        if self.current_question_index >= len(self.evaluation['questions']):
            self.show_results()
            return
        
        question = self.evaluation['questions'][self.current_question_index]
        self.question_label.config(text=question['question'])
        
        # Réinitialiser la sélection
        self.selected_option.set("")
        self.open_answer.set("")

        # Masquer tous les widgets dans la frame de réponses
        for button in self.option_buttons:
            button.pack_forget()

        self.open_answer_label.pack_forget()
        self.open_answer_entry.pack_forget()

        # Afficher les widgets appropriés selon le type de question
        if 'options' in question:
            # Afficher les boutons radio pour les questions à choix multiples
            for idx, option in enumerate("ABCD"):
                if option in question['options']:
                    self.option_buttons[idx].config(text=option + ": " + question['options'][option], state=tk.NORMAL)
                    self.option_buttons[idx].pack(anchor='w')  # Réafficher le bouton
        else:
            # Activer et afficher le champ de réponse ouverte
            self.open_answer_entry.config(state=tk.NORMAL)
            self.open_answer_label.pack(side=tk.LEFT, padx=5)
            self.open_answer_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Activer le bouton "Soumettre"
        self.submit_button.config(state=tk.NORMAL)


    
    def submit_answer(self):
        """Soumet la réponse de l'élève."""
        question = self.evaluation['questions'][self.current_question_index]
        selected_option = self.selected_option.get()
        open_answer = self.open_answer.get()

        if question['type'] == 'open':
            correct_answer = question['answer']
            case_sensitive = question.get('case_sensitive', True)
            ignore_punctuation = question.get('ignore_punctuation', False)
            ignore_spaces = question.get('ignore_spaces', False)
            
            # Gérer la casse
            if not case_sensitive:
                selected_option = selected_option.lower()
                correct_answer = correct_answer.lower()

            # Ignorer la ponctuation
            if ignore_punctuation:
                selected_option = selected_option.translate(str.maketrans('', '', string.punctuation))
                correct_answer = correct_answer.translate(str.maketrans('', '', string.punctuation))

            # Ignorer les espaces
            if ignore_spaces:
                selected_option = selected_option.replace(" ", "")
                correct_answer = correct_answer.replace(" ", "")

            # Validation de la réponse
            if open_answer == correct_answer:
                self.score += 1
        else:
            # Validation des réponses à choix multiple
            if selected_option == question['answer']:
                self.score += 1

        self.current_question_index += 1
        self.show_question()
    
    def show_results(self):
        """Affiche les résultats de l'élève."""
        result_text = f"Évaluation terminée!\nVotre score est {self.score}/{len(self.evaluation['questions'])}"
        messagebox.showinfo("Résultats", result_text)
        self.save_results()

    def save_results(self):
        """Enregistre les résultats de l'élève dans un fichier Excel."""
        os.makedirs(results_dir, exist_ok=True)
        excel_path = os.path.join(results_dir, f"{self.student_name}.xlsx")

        evaluation_id = self.evaluation['id']  # Utiliser l'identifiant de l'évaluation

        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)
        else:
            workbook = Workbook()
            # Ajouter un onglet Résumé des Évaluations
            summary_sheet = workbook.active
            summary_sheet.title = "Résumé des Évaluations"
            summary_sheet.append(["Identifiant", "Matière", "Description", "Date de Passation", "Score"])

        # Ajouter ou mettre à jour l'onglet pour la matière
        subject = self.evaluation.get('subject', 'Général')  # Utiliser 'Général' si la matière n'est pas définie
        if subject not in workbook.sheetnames:
            subject_sheet = workbook.create_sheet(title=subject)
            subject_sheet.append(["Identifiant", "Description", "Date de Passation", "Score"])
        else:
            subject_sheet = workbook[subject]

        # Ajouter les résultats à l'onglet spécifique
        data = [evaluation_id, self.evaluation['title'], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), f"{self.score}/{len(self.evaluation['questions'])}"]
        subject_sheet.append(data)

        # Ajouter les résultats au résumé global
        summary_sheet = workbook["Résumé des Évaluations"]
        summary_data = [evaluation_id, subject, self.evaluation['title'], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), f"{self.score}/{len(self.evaluation['questions'])}"]
        summary_sheet.append(summary_data)

        workbook.save(excel_path)
        
        # Ajoute l'évaluation actuelle à la liste des évaluations complètes
        self.completed_evaluations.add(evaluation_id)
        self.save_completed_evaluations()
        # Met à jour la liste déroulante des évaluations disponibles
        self.reset_interface()
    
    def save_completed_evaluations(self):
        """Enregistre les évaluations complétées dans le fichier JSON."""
        if os.path.exists(completed_evaluations_file):
            with open(completed_evaluations_file, 'r') as f:
                all_evaluations = json.load(f)
        else:
            all_evaluations = {}

        # Utiliser le nom en minuscules pour la clé
        all_evaluations[self.student_name] = list(self.completed_evaluations)
        with open(completed_evaluations_file, 'w') as f:
            json.dump(all_evaluations, f, indent=4)

    
    def reset_interface(self):
        """Réinitialise l'interface après la fin de l'évaluation tout en conservant le prénom de l'élève."""
            
        # Réactive le champ du prénom pour permettre une nouvelle saisie si nécessaire
        self.name_entry.config(state=tk.NORMAL)
        
        # Réinitialise la sélection de l'évaluation
        self.evaluation_var.set("")
        
        # Met à jour la liste des évaluations disponibles en excluant celles déjà passées
        available_titles = [title for title in self.evaluation_mapping if self.evaluation_mapping[title]['id'] not in self.completed_evaluations]
        self.evaluation_menu['values'] = available_titles

        if available_titles:
            # Sélectionner automatiquement la première évaluation disponible
            self.evaluation_menu.current(0)
            self.evaluation_var.set(available_titles[0])
        else:
            messagebox.showinfo("Info", "Aucune évaluation disponible à passer.")
            self.root.quit()

        # Réinitialise l'affichage des questions
        self.question_label.config(text="")
        self.selected_option.set("")
        for button in self.option_buttons:
            button.config(state=tk.DISABLED)
        self.submit_button.config(state=tk.DISABLED)


    def get_current_date(self):
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d")

if __name__ == "__main__":
    root = tk.Tk()
    app = QuizApp(root)
    root.mainloop()

